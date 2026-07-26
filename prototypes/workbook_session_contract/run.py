"""Throwaway terminal prototype for the WorkbookSession contract.

It intentionally uses preset requests instead of an LLM. The point is to feel out the session,
versioning, query, staging, and confirmation state model before production implementation.
"""

from __future__ import annotations

import json
import shutil
import tempfile
import unicodedata
from dataclasses import dataclass, replace
from datetime import date, datetime
from decimal import Decimal
from hashlib import sha256
from numbers import Number
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from state import PendingMutation, SessionState, commit, confirm, reject_unconfirmed_commit, stage


ROOT = Path(__file__).resolve().parents[2]
SOURCES = {
    "listings": ROOT / "docs/task-reqs/Real Estate Listings.xlsx",
    "campaigns": ROOT / "docs/task-reqs/Marketing Campaigns.xlsx",
}
ID_COLUMNS = {"listings": "Listing ID", "campaigns": "Campaign ID"}


def digest(path: Path) -> str:
    return sha256(path.read_bytes()).hexdigest()[:12]


def workbook_fingerprint(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    return {
        "sheets": workbook.sheetnames,
        "dimensions": {sheet.title: sheet.max_row * sheet.max_column for sheet in workbook},
        "headers": {sheet.title: [cell.value for cell in next(sheet.iter_rows(max_row=1))] for sheet in workbook},
        "formula_cells": sum(
            cell.data_type == "f" for sheet in workbook for row in sheet.iter_rows() for cell in row
        ),
        "styled_cells": sum(
            cell.has_style for sheet in workbook for row in sheet.iter_rows() for cell in row
        ),
    }


@dataclass
class StagedChange:
    operation: str
    workbook: str
    before: dict[str, object] | None
    after: dict[str, object] | None
    target_id: str
    source_hash: str
    expected_changes: tuple["ExpectedChange", ...]


@dataclass(frozen=True)
class ExpectedChange:
    sheet: str
    cell: str
    value: object
    value_type: str | None = None


def normalized_value(value: object) -> tuple[str, str | None]:
    """Normalize values for semantic comparison, not binary workbook equality."""
    if value is None or (not isinstance(value, str) and bool(pd.isna(value))):
        return ("null", None)
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return ("date", value.isoformat())
    if isinstance(value, Number) and not isinstance(value, bool):
        return ("number", str(Decimal(str(value)).normalize()))
    if isinstance(value, str):
        return ("string", unicodedata.normalize("NFC", value).replace("\r\n", "\n"))
    return (type(value).__name__, repr(value))


def change_value(value: object, value_type: str | None = None) -> dict[str, object]:
    normalized_type, normalized = normalized_value(value)
    return {
        "repr": repr(value),
        "value_type": value_type or type(value).__name__,
        "normalized_type": normalized_type,
        "normalized_value": normalized,
    }


class WorkbookSessionPrototype:
    """Minimal, deliberately non-production session implementation."""

    def __init__(self) -> None:
        self.root = Path(tempfile.mkdtemp(prefix="workbook-session-prototype-"))
        self.state = SessionState()
        self.staged: StagedChange | None = None
        self.last_result: dict[str, object] = {"status": "ok", "message": "Session created"}
        self.sources = {name: path for name, path in SOURCES.items()}
        self.versions: dict[str, list[Path]] = {name: [] for name in SOURCES}
        self.compatible: dict[str, bool] = {}
        for name, source in self.sources.items():
            initial = self.root / f"{name}-v1.xlsx"
            shutil.copy2(source, initial)
            source_fingerprint = workbook_fingerprint(source)
            workbook = load_workbook(initial)
            workbook.save(initial)
            self.compatible[name] = source_fingerprint == workbook_fingerprint(initial)
            self.versions[name].append(initial)

    def active_path(self, workbook: str) -> Path:
        return self.versions[workbook][-1]

    def describe(self, workbook: str) -> dict[str, object]:
        frame = pd.read_excel(self.active_path(workbook))
        return {"status": "ok", "workbook": workbook, "rows": len(frame), "columns": list(frame.columns)}

    def query(self, workbook: str, *, field: str, value: object) -> dict[str, object]:
        frame = pd.read_excel(self.active_path(workbook))
        if field not in frame.columns:
            return {"status": "rejected", "error": "unknown_field", "field": field}
        matches = frame.loc[frame[field] == value]
        return {
            "status": "ok",
            "total_matches": len(matches),
            "truncated": len(matches) > 5,
            "rows": matches.head(5).to_dict(orient="records"),
        }

    def campaign_metric(self, metric: str) -> dict[str, object]:
        frame = pd.read_excel(self.active_path("campaigns"))
        definitions = {
            "ctr": ("Clicks", "Impressions", "sum(Clicks) / sum(Impressions)"),
            "cpa": ("Amount Spent", "Conversions", "sum(Amount Spent) / sum(Conversions)"),
            "roas": ("Revenue Generated", "Amount Spent", "sum(Revenue Generated) / sum(Amount Spent)"),
        }
        numerator, denominator, calculation = definitions[metric]
        denominator_value = float(frame[denominator].sum())
        return {
            "status": "unavailable" if denominator_value == 0 else "ok",
            "metric": metric,
            "calculation": calculation,
            "calculation_source": "tool_computed",
            "value": None if denominator_value == 0 else float(frame[numerator].sum()) / denominator_value,
        }

    def contract_assertions(self) -> dict[str, object]:
        metrics = [self.campaign_metric(metric) for metric in ("ctr", "cpa", "roas")]
        source_check = all(item["calculation_source"] == "tool_computed" for item in metrics)
        field_check = all("formula" not in item for item in metrics)

        probe = self.root / "formula-preservation-probe.xlsx"
        workbook = load_workbook(self.active_path("listings"))
        sheet = workbook.active
        probe_column = sheet.max_column + 1
        probe_cell = sheet.cell(2, probe_column)
        probe_cell.value = "=1+1"
        workbook.save(probe)
        before = load_workbook(probe, data_only=False).active.cell(2, probe_column).value
        workbook = load_workbook(probe, data_only=False)
        workbook.active.cell(3, probe_column).value = "preserved without evaluation"
        workbook.save(probe)
        expression = load_workbook(probe, data_only=False).active.cell(2, probe_column).value
        cached_value = load_workbook(probe, data_only=True).active.cell(2, probe_column).value
        formula_check = before == expression == "=1+1" and cached_value is None

        checks = {
            "tool_computed_ctr_cpa_roas": source_check,
            "metric_responses_exclude_formula": field_check,
            "spreadsheet_formula_preserved_not_evaluated": formula_check,
        }
        return {
            "status": "ok" if all(checks.values()) else "rejected",
            "checks": checks,
            "spreadsheet_formula": {
                "expression": expression,
                "cached_value": cached_value,
                "calculation_source": "spreadsheet_formula",
            },
        }

    def aurora_ambiguity(self) -> dict[str, object]:
        frame = pd.read_excel(self.active_path("listings"))
        states = sorted(frame.loc[frame["City"] == "Aurora", "State"].dropna().unique().tolist())
        return {
            "status": "needs_clarification",
            "error": "ambiguous_city_scope",
            "candidates": states,
        }

    def stage_update(self) -> None:
        workbook = "listings"
        frame = pd.read_excel(self.active_path(workbook))
        row = frame.iloc[0]
        target_id = str(row[ID_COLUMNS[workbook]])
        before = {"List Price": row["List Price"]}
        after = {"List Price": int(row["List Price"]) + 1}
        self._stage("update", workbook, before, after, target_id)

    def stage_insert(self) -> None:
        workbook = "listings"
        frame = pd.read_excel(self.active_path(workbook))
        target_id = "PROTO-LISTING-001"
        if target_id in set(frame[ID_COLUMNS[workbook]].astype(str)):
            target_id = f"PROTO-LISTING-{self.state.active_version + 1:03d}"
        after = {
            "Listing ID": target_id,
            "Property Type": "Prototype",
            "City": "Phoenix",
            "State": "AZ",
            "Bedrooms": 1,
            "Bathrooms": 1,
            "Square Footage": 500,
            "Year Built": 2026,
            "List Price": 100000,
            "Sale Price": None,
            "Listing Status": "Active",
        }
        self._stage("insert", workbook, None, after, target_id)

    def stage_delete(self) -> None:
        workbook = "listings"
        frame = pd.read_excel(self.active_path(workbook))
        row = frame.iloc[1]
        target_id = str(row[ID_COLUMNS[workbook]])
        self._stage("delete", workbook, row.to_dict(), None, target_id)

    def _stage(
        self,
        operation: str,
        workbook: str,
        before: dict[str, object] | None,
        after: dict[str, object] | None,
        target_id: str,
    ) -> None:
        if not self.compatible[workbook]:
            self.state = replace(self.state, last_status="Rejected: workbook failed no-op compatibility check")
            return
        worksheet = load_workbook(self.active_path(workbook), data_only=False).active
        headers = {cell.value: cell.column for cell in next(worksheet.iter_rows(max_row=1))}
        if operation == "delete":
            expected_changes = (
                ExpectedChange(worksheet.title, f"record:{ID_COLUMNS[workbook]}={target_id}", None, "record_absence"),
            )
        else:
            target_row = worksheet.max_row + 1 if operation == "insert" else next(
                row
                for row in range(2, worksheet.max_row + 1)
                if str(worksheet.cell(row, headers[ID_COLUMNS[workbook]]).value) == target_id
            )
            expected_changes = tuple(
                ExpectedChange(worksheet.title, f"{get_column_letter(headers[column])}{target_row}", value)
                for column, value in (after or {}).items()
            )
        change = StagedChange(
            operation,
            workbook,
            before,
            after,
            target_id,
            digest(self.active_path(workbook)),
            expected_changes,
        )
        self.staged = change
        self.state = stage(
            self.state,
            PendingMutation(change.operation, change.workbook, f"{change.operation} {change.target_id}"),
        )

    def record_confirmation(self) -> None:
        self.state = confirm(self.state)

    def commit(self) -> None:
        if self.staged is None or not self.state.confirmed:
            self.state = reject_unconfirmed_commit(self.state)
            self.last_result = {"status": "rejected", "error": "confirmation_required"}
            return
        change = self.staged
        if digest(self.active_path(change.workbook)) != change.source_hash:
            self._reject_transaction(change, "source_version_changed", [])
            return
        next_path = self.root / f"{change.workbook}-v{self.state.active_version + 1}.xlsx"
        try:
            shutil.copy2(self.active_path(change.workbook), next_path)
            self._apply_candidate(change, next_path)
            verified, expected_changes, actual_changes = self._verify_candidate(change, next_path)
        except (RuntimeError, StopIteration, ValueError) as error:
            next_path.unlink(missing_ok=True)
            self._reject_transaction(change, f"candidate_write_failed: {error}", [])
            return
        if not verified:
            next_path.unlink(missing_ok=True)
            self._reject_transaction(change, "semantic_diff_mismatch", actual_changes, expected_changes)
            return
        self.versions[change.workbook].append(next_path)
        if len(self.versions[change.workbook]) > 5:
            evicted = self.versions[change.workbook].pop(0)
            evicted.unlink(missing_ok=True)
        self.state = commit(self.state)
        self.staged = None
        self.last_result = {
            "status": "ok",
            "candidate_version": self.state.active_version,
            "expected_changes": expected_changes,
            "actual_changes": actual_changes,
        }

    def _apply_candidate(self, change: StagedChange, candidate: Path) -> None:
        workbook = load_workbook(candidate, data_only=False)
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        id_column = headers.index(ID_COLUMNS[change.workbook]) + 1
        matching_rows = [row for row in range(2, sheet.max_row + 1) if str(sheet.cell(row, id_column).value) == change.target_id]
        if change.operation == "insert":
            for expected in change.expected_changes:
                sheet[expected.cell].value = expected.value
        elif len(matching_rows) != 1:
            raise RuntimeError(f"Expected one Stable ID match; found {len(matching_rows)}")
        elif change.operation == "update":
            for column, value in (change.after or {}).items():
                sheet.cell(matching_rows[0], headers.index(column) + 1).value = value
        else:
            sheet.delete_rows(matching_rows[0], 1)
        workbook.save(candidate)

    def _verify_candidate(
        self, change: StagedChange, candidate: Path
    ) -> tuple[bool, list[dict[str, object]], list[dict[str, object]]]:
        workbook = load_workbook(candidate, data_only=False)
        sheet = workbook.active
        headers = {cell.value: cell.column for cell in next(sheet.iter_rows(max_row=1))}
        id_column = headers[ID_COLUMNS[change.workbook]]
        expected_payload: list[dict[str, object]] = []
        actual_payload: list[dict[str, object]] = []
        matches = [
            row
            for row in range(2, sheet.max_row + 1)
            if str(sheet.cell(row, id_column).value) == change.target_id
        ]
        for expected in change.expected_changes:
            expected_item = {"sheet": expected.sheet, "cell": expected.cell, **change_value(expected.value, expected.value_type)}
            if expected.value_type == "record_absence":
                actual_value: object = None if not matches else "present"
                actual_type = "record_absence" if not matches else "record_presence"
            else:
                actual_value = sheet[expected.cell].value
                actual_type = None
            actual_item = {"sheet": expected.sheet, "cell": expected.cell, **change_value(actual_value, actual_type)}
            expected_payload.append(expected_item)
            actual_payload.append(actual_item)
        matches_expected = all(
            expected["normalized_type"] == actual["normalized_type"]
            and expected["normalized_value"] == actual["normalized_value"]
            for expected, actual in zip(expected_payload, actual_payload, strict=True)
        )
        return matches_expected, expected_payload, actual_payload

    def _reject_transaction(
        self,
        change: StagedChange,
        error: str,
        actual_changes: list[dict[str, object]],
        expected_changes: list[dict[str, object]] | None = None,
    ) -> None:
        expected = expected_changes or [
            {"sheet": item.sheet, "cell": item.cell, **change_value(item.value, item.value_type)}
            for item in change.expected_changes
        ]
        self.state = replace(self.state, last_status=f"Rejected transaction: {error}")
        self.last_result = {
            "status": "rejected",
            "error": error,
            "active_version": self.state.active_version,
            "expected_changes": expected,
            "actual_changes": actual_changes,
        }

    def snapshot(self) -> dict[str, object]:
        return {
            "temporary_root": str(self.root),
            "active_version": self.state.active_version,
            "retained_versions": self.state.retained_versions,
            "compatibility": self.compatible,
            "active_hashes": {name: digest(self.active_path(name)) for name in SOURCES},
            "pending": None if self.state.pending is None else self.state.pending.summary,
            "confirmed": self.state.confirmed,
            "last_status": self.state.last_status,
            "last_result": self.last_result,
        }


def frame(session: WorkbookSessionPrototype) -> None:
    print("\033[2J\033[H", end="")
    print("\033[1mWorkbookSession contract prototype — THROWAWAY\033[0m")
    print(json.dumps(session.snapshot(), indent=2, default=str))
    print("\n\033[1mKeyboard shortcuts\033[0m")
    print("[d] describe both   [p] property query   [k] campaign KPI   [v] contract assertions")
    print("[a] Aurora ambiguity")
    print("[f] invalid field   [i] stage insert     [u] stage update   [x] stage delete")
    print("[y] confirm staged  [c] commit staged    [q] quit")


def main() -> None:
    session = WorkbookSessionPrototype()
    while True:
        frame(session)
        key = input("\nAction: ").strip().lower()
        if key == "q":
            return
        if key == "d":
            session.state = replace(session.state, last_status=json.dumps([session.describe(name) for name in SOURCES]))
        elif key == "p":
            session.state = replace(
                session.state,
                last_status=json.dumps(session.query("listings", field="Listing Status", value="Active")),
            )
        elif key == "k":
            session.state = replace(session.state, last_status=json.dumps(session.campaign_metric("ctr")))
        elif key == "v":
            session.state = replace(session.state, last_status=json.dumps(session.contract_assertions()))
        elif key == "a":
            session.state = replace(session.state, last_status=json.dumps(session.aurora_ambiguity()))
        elif key == "f":
            session.state = replace(
                session.state,
                last_status=json.dumps(session.query("listings", field="Unsafe Field", value="x")),
            )
        elif key == "i":
            session.stage_insert()
        elif key == "u":
            session.stage_update()
        elif key == "x":
            session.stage_delete()
        elif key == "y":
            session.record_confirmation()
        elif key == "c":
            session.commit()
        else:
            session.state = replace(session.state, last_status=f"Unknown action: {key}")


if __name__ == "__main__":
    main()
