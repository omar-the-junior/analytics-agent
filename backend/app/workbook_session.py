"""Bounded, transactional workbook operations for one Session Workbook."""

from __future__ import annotations

import hashlib
import shutil
import tempfile
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal

import pandas as pd
from openpyxl import load_workbook
from pydantic import (
    BaseModel,
    ConfigDict,
    Field,
    StrictInt,
    StrictStr,
    TypeAdapter,
    ValidationError,
    model_validator,
)

from app.agent_loop import ToolCall, ToolResult

ROOT = Path(__file__).resolve().parents[2]
SOURCES = {
    "listings": ROOT / "docs" / "task-reqs" / "Real Estate Listings.xlsx",
    "campaigns": ROOT / "docs" / "task-reqs" / "Marketing Campaigns.xlsx",
}
ID_COLUMNS = {"listings": "Listing ID", "campaigns": "Campaign ID"}
MAX_ROWS = 100

FilterValue = str | int | float | bool | None
JsonValue = str | int | float | bool | None


class QueryFilter(BaseModel):
    """One allowlisted predicate against the bound Session Workbook."""

    model_config = ConfigDict(extra="forbid")

    column: StrictStr = Field(min_length=1)
    operator: Literal[
        "eq", "in", "lt", "lte", "gt", "gte", "between", "is_null", "not_null", "overlaps"
    ] = "eq"
    value: Any = None


class OrderBy(BaseModel):
    model_config = ConfigDict(extra="forbid")

    column: StrictStr = Field(min_length=1)
    direction: Literal["asc", "desc"] = "asc"


class CalculationRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")

    kind: Literal["rows", "count", "sum", "min", "max"] = "rows"
    column: StrictStr | None = None

    @model_validator(mode="after")
    def require_metric_column(self) -> CalculationRequest:
        if self.kind in {"sum", "min", "max"} and self.column is None:
            raise ValueError("metric calculation requires a column")
        if self.kind in {"rows", "count"} and self.column is not None:
            raise ValueError("calculation does not accept a column")
        return self


class TableQueryResult(BaseModel):
    """Normalized row result. ``row_count`` is before the display limit."""

    model_config = ConfigDict(extra="forbid")

    kind: Literal["table"] = "table"
    columns: list[str]
    rows: list[list[JsonValue]]
    row_count: int = Field(ge=0)
    truncated: bool
    stable_id_field: str
    calculation_source: Literal["tool_computed"] = "tool_computed"


class MetricQueryResult(BaseModel):
    model_config = ConfigDict(extra="forbid")

    kind: Literal["metric"] = "metric"
    metric: Literal["count", "sum", "min", "max"]
    value: JsonValue = None
    column: str | None = None
    row_count: int = Field(ge=0)
    unavailable: bool = False
    reason: str | None = None
    calculation_source: Literal["tool_computed"] = "tool_computed"


class SelectionQueryResult(BaseModel):
    model_config = ConfigDict(extra="forbid")

    kind: Literal["selection"] = "selection"
    column: str
    value: JsonValue
    row: dict[str, JsonValue]
    stable_id_field: str
    stable_id: str
    calculation_source: Literal["tool_computed"] = "tool_computed"


QueryResult = TableQueryResult | MetricQueryResult | SelectionQueryResult
QUERY_RESULT_ADAPTER = TypeAdapter(QueryResult)


class QueryRequest(BaseModel):
    """The complete, bounded representation accepted by ``query_workbook``."""

    model_config = ConfigDict(extra="forbid")

    # The mapping form and aggregate/column fields are accepted only as a migration path for
    # existing model transcripts. New callers send a list of predicates and ``calculation``.
    filters: list[QueryFilter] | dict[str, FilterValue] = Field(default_factory=list)
    select: list[StrictStr] | None = None
    order_by: list[OrderBy] = Field(default_factory=list, max_length=25)
    calculation: CalculationRequest | None = None
    aggregate: Literal["rows", "count", "sum", "min", "max"] | None = None
    column: str | None = None
    limit: StrictInt = Field(default=10, ge=1, le=MAX_ROWS)
    presentation: Literal["table"] = "table"

    @model_validator(mode="after")
    def validate_migration_shape(self) -> QueryRequest:
        if self.calculation is not None and self.aggregate is not None:
            raise ValueError("use calculation or aggregate, not both")
        if self.calculation is not None and self.column is not None:
            raise ValueError("calculation column must be nested")
        if self.calculation is None and self.aggregate is None and self.column is not None:
            raise ValueError("column requires a calculation")
        return self

    @property
    def resolved_calculation(self) -> CalculationRequest:
        if self.calculation is not None:
            return self.calculation
        return CalculationRequest(kind=self.aggregate or "rows", column=self.column)


class MutationRequest(BaseModel):
    """The complete, bounded representation accepted by ``stage_mutation``."""

    model_config = ConfigDict(extra="forbid")

    operation: Literal["insert", "update", "delete"]
    target_id: StrictStr = Field(min_length=1, pattern=r"\S")
    values: dict[str, FilterValue] = Field(default_factory=dict)


def _digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _value(value: Any) -> Any:
    """Convert pandas scalars into JSON-safe, user-facing values."""
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.date().isoformat()
    return value.item() if hasattr(value, "item") else value


@dataclass(frozen=True)
class StagedMutation:
    stage_id: str
    operation: str
    target_id: str
    values: dict[str, Any]
    source_hash: str
    before: dict[str, Any] | None
    after: dict[str, Any] | None
    warnings: list[str]


class WorkbookSession:
    """A chat-scoped service with no authority outside one supplied workbook."""

    def __init__(self, workbook: str, workspace: Path | None = None) -> None:
        if workbook not in SOURCES:
            raise ValueError("unknown_workbook")
        self.workbook = workbook
        self._source = SOURCES[workbook]
        self._root = workspace or Path(tempfile.mkdtemp(prefix="workbook-session-"))
        self._root.mkdir(parents=True, exist_ok=True)
        self._active = self._root / f"{workbook}-v1.xlsx"
        shutil.copy2(self._source, self._active)
        self._source_identity = _digest(self._active)
        self._version = 1
        self._staged: StagedMutation | None = None

    @property
    def source_hash(self) -> str:
        return self._source_identity

    @property
    def active_path(self) -> Path:
        return self._active

    @property
    def pending_stage(self) -> dict[str, Any] | None:
        stage = self._staged
        if stage is None:
            return None
        identifier = ID_COLUMNS[self.workbook]
        payload: dict[str, Any] = {
            "stage_id": stage.stage_id,
            "operation": stage.operation,
            "stable_id_field": identifier,
            "stable_id": stage.target_id,
            "warnings": stage.warnings,
        }
        if stage.operation == "update":
            payload["preview"] = {
                "kind": "field_diff",
                "columns": ["Field", "Before", "After"],
                "rows": [
                    [column, stage.before[column], stage.after[column]]
                    for column in stage.values
                    if stage.before is not None
                    and stage.after is not None
                    and stage.before[column] != stage.after[column]
                ],
            }
        elif stage.operation == "insert":
            payload["preview"] = {
                "kind": "after_row",
                "columns": list(stage.after or {}),
                "rows": [list((stage.after or {}).values())],
            }
        else:
            payload["preview"] = {
                "kind": "before_row",
                "columns": list(stage.before or {}),
                "rows": [list((stage.before or {}).values())],
            }
        return payload

    def discard_stage(self) -> None:
        self._staged = None

    def _frame(self) -> pd.DataFrame:
        return pd.read_excel(self._active)

    def describe_workbook(self) -> ToolResult:
        frame = self._frame()
        return ToolResult(
            status="ok",
            payload={
                "workbook": self.workbook,
                "rows": len(frame),
                "columns": list(frame.columns),
                "stable_id": ID_COLUMNS[self.workbook],
                "source_identity": self._source_identity,
                "version": self._version,
            },
        )

    def _query_filters(self, query: QueryRequest) -> list[QueryFilter]:
        if isinstance(query.filters, dict):
            return [
                QueryFilter(column=column, operator="eq", value=value)
                for column, value in query.filters.items()
            ]
        return query.filters

    @staticmethod
    def _range_value(series: pd.Series, value: Any) -> float | pd.Timestamp | None:
        if pd.api.types.is_numeric_dtype(series):
            if isinstance(value, bool) or not isinstance(value, int | float):
                return None
            return float(value)
        if pd.api.types.is_datetime64_any_dtype(series):
            if not isinstance(value, str):
                return None
            converted = pd.to_datetime(value, errors="coerce")
            return converted if not pd.isna(converted) else None
        return None

    @staticmethod
    def _is_exact_value_compatible(series: pd.Series, value: Any) -> bool:
        """Keep empty results distinct from an operand of the wrong workbook type."""
        if value is None:
            return True
        if pd.api.types.is_numeric_dtype(series):
            return not isinstance(value, bool) and isinstance(value, int | float)
        if pd.api.types.is_datetime64_any_dtype(series):
            return isinstance(value, str) and not pd.isna(pd.to_datetime(value, errors="coerce"))
        return isinstance(value, str | bool)

    def _apply_filters(
        self, frame: pd.DataFrame, filters: list[QueryFilter]
    ) -> pd.DataFrame | ToolResult:
        known_columns = set(frame.columns)
        invalid_columns = sorted(
            {
                filter_.column
                for filter_ in filters
                if filter_.column not in known_columns
                and not (filter_.operator == "overlaps" and filter_.column == "Campaign Interval")
            }
        )
        if invalid_columns:
            return ToolResult(
                status="rejected",
                payload={"error_code": "unknown_field", "fields": invalid_columns},
            )

        for filter_ in filters:
            if filter_.operator == "eq" and not self._is_exact_value_compatible(
                frame[filter_.column], filter_.value
            ):
                return ToolResult(status="rejected", payload={"error_code": "invalid_filter_value"})
            if filter_.operator == "in":
                if not isinstance(filter_.value, list) or not filter_.value or any(
                    not self._is_exact_value_compatible(frame[filter_.column], value)
                    for value in filter_.value
                ):
                    return ToolResult(
                        status="rejected", payload={"error_code": "invalid_filter_value"}
                    )

        city_values = [
            filter_.value
            for filter_ in filters
            if filter_.column == "City" and filter_.operator == "eq"
        ]
        city_values.extend(
            value
            for filter_ in filters
            if filter_.column == "City" and filter_.operator == "in"
            for value in filter_.value
        )
        has_state_scope = any(filter_.column == "State" for filter_ in filters)
        if self.workbook == "listings" and city_values and not has_state_scope:
            for city in city_values:
                states = sorted(
                    str(value)
                    for value in frame.loc[frame["City"] == city, "State"].dropna().unique()
                )
                if len(states) > 1:
                    return ToolResult(
                        status="needs_clarification",
                        payload={"error_code": "ambiguous_city_scope", "candidates": states},
                    )

        filtered = frame
        for filter_ in filters:
            column = filter_.column
            operator = filter_.operator
            value = filter_.value
            if operator == "overlaps":
                if self.workbook != "campaigns" or column != "Campaign Interval":
                    return ToolResult(status="rejected", payload={"error_code": "invalid_filter"})
                if not isinstance(value, dict) or set(value) != {"start", "end"}:
                    return ToolResult(
                        status="rejected", payload={"error_code": "invalid_filter_value"}
                    )
                start = pd.to_datetime(value["start"], errors="coerce")
                end = pd.to_datetime(value["end"], errors="coerce")
                if pd.isna(start) or pd.isna(end) or start > end:
                    return ToolResult(
                        status="rejected", payload={"error_code": "invalid_filter_value"}
                    )
                filtered = filtered.loc[
                    (filtered["Start Date"] <= end) & (filtered["End Date"] >= start)
                ]
                continue

            series = filtered[column]
            if operator == "eq":
                filtered = filtered.loc[series.isna() if value is None else series == value]
            elif operator == "in":
                filtered = filtered.loc[series.isin(value)]
            elif operator in {"is_null", "not_null"}:
                if value is not None:
                    return ToolResult(
                        status="rejected", payload={"error_code": "invalid_filter_value"}
                    )
                filtered = filtered.loc[series.isna() if operator == "is_null" else series.notna()]
            elif operator in {"lt", "lte", "gt", "gte"}:
                comparable = self._range_value(series, value)
                if comparable is None:
                    return ToolResult(
                        status="rejected", payload={"error_code": "invalid_filter_value"}
                    )
                comparisons = {
                    "lt": series < comparable,
                    "lte": series <= comparable,
                    "gt": series > comparable,
                    "gte": series >= comparable,
                }
                filtered = filtered.loc[comparisons[operator]]
            elif operator == "between":
                if not isinstance(value, list) or len(value) != 2:
                    return ToolResult(
                        status="rejected", payload={"error_code": "invalid_filter_value"}
                    )
                lower = self._range_value(series, value[0])
                upper = self._range_value(series, value[1])
                if lower is None or upper is None or lower > upper:
                    return ToolResult(
                        status="rejected", payload={"error_code": "invalid_filter_value"}
                    )
                filtered = filtered.loc[(series >= lower) & (series <= upper)]
            else:
                return ToolResult(status="rejected", payload={"error_code": "invalid_filter"})
        return filtered

    def _ordered_rows(
        self, frame: pd.DataFrame, order_by: list[OrderBy]
    ) -> pd.DataFrame | ToolResult:
        unknown = sorted({order.column for order in order_by} - set(frame.columns))
        if unknown:
            return ToolResult(
                status="rejected", payload={"error_code": "unknown_field", "fields": unknown}
            )
        if len({order.column for order in order_by}) != len(order_by):
            return ToolResult(status="rejected", payload={"error_code": "duplicate_order_column"})
        if not order_by:
            return frame
        identifier = ID_COLUMNS[self.workbook]
        columns = [order.column for order in order_by]
        ascending = [order.direction == "asc" for order in order_by]
        if identifier not in columns:
            columns.append(identifier)
            ascending.append(True)
        return frame.sort_values(columns, ascending=ascending, kind="stable", na_position="last")

    def _table_result(self, frame: pd.DataFrame, query: QueryRequest) -> ToolResult:
        ordered = self._ordered_rows(frame, query.order_by)
        if isinstance(ordered, ToolResult):
            return ordered
        identifier = ID_COLUMNS[self.workbook]
        selected_columns = list(query.select) if query.select is not None else list(frame.columns)
        unknown = sorted(set(selected_columns) - set(frame.columns))
        if unknown:
            return ToolResult(
                status="rejected", payload={"error_code": "unknown_field", "fields": unknown}
            )
        if len(set(selected_columns)) != len(selected_columns):
            return ToolResult(status="rejected", payload={"error_code": "duplicate_select_column"})
        if identifier not in selected_columns:
            selected_columns.append(identifier)
        result = TableQueryResult(
            columns=selected_columns,
            rows=[
                [_value(value) for value in row]
                for row in ordered.loc[:, selected_columns]
                .head(query.limit)
                .itertuples(index=False, name=None)
            ],
            row_count=len(frame),
            truncated=len(frame) > query.limit,
            stable_id_field=identifier,
        )
        return ToolResult(status="ok", payload=result.model_dump())

    def query_workbook(self, arguments: Any) -> ToolResult:
        try:
            query = QueryRequest.model_validate(arguments, strict=True)
        except ValidationError:
            return ToolResult(status="rejected", payload={"error_code": "invalid_query"})
        frame = self._frame()
        filtered = self._apply_filters(frame, self._query_filters(query))
        if isinstance(filtered, ToolResult):
            return filtered
        calculation = query.resolved_calculation
        if calculation.kind == "rows":
            return self._table_result(filtered, query)
        if calculation.kind == "count":
            return ToolResult(
                status="ok",
                payload=MetricQueryResult(
                    metric="count", value=len(filtered), row_count=len(filtered)
                ).model_dump(),
            )
        column = calculation.column
        if column not in filtered.columns or not pd.api.types.is_numeric_dtype(filtered[column]):
            return ToolResult(status="rejected", payload={"error_code": "invalid_aggregate_column"})
        values = filtered[column].dropna()
        if values.empty:
            return ToolResult(
                status="ok",
                payload=MetricQueryResult(
                    metric=calculation.kind,
                    column=column,
                    row_count=len(filtered),
                    unavailable=True,
                    reason="no_non_null_values",
                ).model_dump(),
            )
        if calculation.kind == "sum":
            return ToolResult(
                status="ok",
                payload=MetricQueryResult(
                    metric="sum", column=column, value=_value(values.sum()), row_count=len(filtered)
                ).model_dump(),
            )
        candidates = filtered.dropna(subset=[column])
        identifier = ID_COLUMNS[self.workbook]
        selected = candidates.sort_values(
            [column, identifier], ascending=[calculation.kind == "min", True], kind="stable"
        ).iloc[0]
        row = {key: _value(value) for key, value in selected.to_dict().items()}
        return ToolResult(
            status="ok",
            payload=SelectionQueryResult(
                column=column,
                value=row[column],
                row=row,
                stable_id_field=identifier,
                stable_id=row[identifier],
            ).model_dump(),
        )

    def _validate_mutation_values(
        self, frame: pd.DataFrame, values: dict[str, FilterValue]
    ) -> ToolResult | None:
        invalid = []
        for column, value in values.items():
            series = frame[column]
            if pd.api.types.is_numeric_dtype(series) and (
                isinstance(value, bool) or not isinstance(value, int | float)
            ):
                invalid.append(column)
            elif pd.api.types.is_datetime64_any_dtype(series) and (
                not isinstance(value, str) or pd.isna(pd.to_datetime(value, errors="coerce"))
            ):
                invalid.append(column)
        if invalid:
            return ToolResult(
                status="rejected", payload={"error_code": "invalid_field_value", "fields": invalid}
            )
        if self.workbook == "listings" and "Listing Status" in values:
            statuses = set(frame["Listing Status"].dropna().unique())
            if values["Listing Status"] not in statuses:
                return ToolResult(
                    status="rejected", payload={"error_code": "invalid_listing_status"}
                )
        return None

    def stage_mutation(self, arguments: Any) -> ToolResult:
        try:
            mutation = MutationRequest.model_validate(arguments, strict=True)
        except ValidationError:
            return ToolResult(status="rejected", payload={"error_code": "invalid_mutation"})
        operation = mutation.operation
        target_id = mutation.target_id
        values = mutation.values
        if operation == "delete" and values:
            return ToolResult(status="rejected", payload={"error_code": "invalid_delete_values"})
        frame = self._frame()
        identifier = ID_COLUMNS[self.workbook]
        if set(values) - set(frame.columns):
            return ToolResult(status="rejected", payload={"error_code": "unknown_field"})
        if operation == "update" and identifier in values:
            return ToolResult(status="rejected", payload={"error_code": "stable_id_immutable"})
        invalid_values = self._validate_mutation_values(frame, values)
        if invalid_values:
            return invalid_values
        matches = frame.loc[frame[identifier].astype(str) == target_id]
        if operation == "insert":
            if (
                not values
                or str(values.get(identifier, target_id)) != target_id
                or not matches.empty
            ):
                return ToolResult(
                    status="rejected", payload={"error_code": "invalid_insert_target"}
                )
            missing = set(frame.columns) - set(values)
            if missing:
                return ToolResult(
                    status="rejected",
                    payload={"error_code": "missing_insert_fields", "fields": sorted(missing)},
                )
            before, after = None, {key: _value(value) for key, value in values.items()}
        else:
            if len(matches) != 1:
                return ToolResult(status="rejected", payload={"error_code": "stable_id_not_found"})
            before = {key: _value(value) for key, value in matches.iloc[0].to_dict().items()}
            after = (
                None
                if operation == "delete"
                else {**before, **{key: _value(value) for key, value in values.items()}}
            )
        if self.workbook == "listings" and after is not None:
            if after.get("Listing Status") == "Sold" and after.get("Sale Price") is None:
                return ToolResult(status="rejected", payload={"error_code": "missing_sale_price"})
        warnings = []
        if self.workbook == "listings" and after is not None:
            if after.get("Listing Status") == "Active" and after.get("Sale Price") is not None:
                warnings.append("active_listing_retains_sale_price")
        stage_id = hashlib.sha256(
            f"{self._version}:{operation}:{target_id}:{values}".encode()
        ).hexdigest()[:16]
        self._staged = StagedMutation(
            stage_id, operation, target_id, values, _digest(self._active), before, after, warnings
        )
        return ToolResult(
            status="ok",
            payload={"status": "confirmation_required"} | (self.pending_stage or {}),
        )

    def commit_mutation(self, arguments: dict[str, Any]) -> ToolResult:
        stage = self._staged
        if stage is None or arguments != {"stage_id": stage.stage_id}:
            return ToolResult(status="rejected", payload={"error_code": "confirmation_required"})
        if _digest(self._active) != stage.source_hash:
            return ToolResult(status="rejected", payload={"error_code": "stale_stage"})
        candidate = self._root / f"{self.workbook}-v{self._version + 1}.xlsx"
        shutil.copy2(self._active, candidate)
        workbook = load_workbook(candidate)
        sheet = workbook.active
        headers = {cell.value: cell.column for cell in sheet[1]}
        identifier = ID_COLUMNS[self.workbook]
        row_numbers = [
            row
            for row in range(2, sheet.max_row + 1)
            if str(sheet.cell(row, headers[identifier]).value) == stage.target_id
        ]
        if stage.operation == "insert":
            row = sheet.max_row + 1
            for column, value in (stage.after or {}).items():
                sheet.cell(row, headers[column]).value = value
        elif stage.operation == "update" and len(row_numbers) == 1:
            for column, value in stage.values.items():
                sheet.cell(row_numbers[0], headers[column]).value = value
        elif stage.operation == "delete" and len(row_numbers) == 1:
            sheet.delete_rows(row_numbers[0])
        else:
            candidate.unlink(missing_ok=True)
            return ToolResult(status="rejected", payload={"error_code": "stable_id_not_found"})
        workbook.save(candidate)
        verification = self._verify(stage, candidate)
        if verification is None:
            candidate.unlink(missing_ok=True)
            return ToolResult(
                status="rejected", payload={"error_code": "artifact_verification_failed"}
            )
        self._active = candidate
        self._version += 1
        self._staged = None
        return ToolResult(
            status="ok",
            payload={
                "artifact": candidate.name,
                "version": self._version,
                "verified": True,
                "stable_id": stage.target_id,
                "verification": verification,
            },
        )

    def _verify(self, stage: StagedMutation, candidate: Path) -> dict[str, int] | None:
        """Reopen an artifact and prove the authorized diff is its only data change."""
        before_book = load_workbook(self._active, data_only=False)
        candidate_book = load_workbook(candidate, data_only=False)
        before_sheet = before_book.active
        candidate_sheet = candidate_book.active
        before_headers = {cell.value: cell.column for cell in before_sheet[1]}
        candidate_headers = {cell.value: cell.column for cell in candidate_sheet[1]}
        identifier = ID_COLUMNS[self.workbook]
        if before_headers != candidate_headers or identifier not in before_headers:
            return None

        def rows_by_id(sheet: Any, headers: dict[Any, int]) -> dict[str, dict[Any, Any]]:
            return {
                str(sheet.cell(row, headers[identifier]).value): {
                    name: sheet.cell(row, column).value for name, column in headers.items()
                }
                for row in range(2, sheet.max_row + 1)
            }

        before_rows = rows_by_id(before_sheet, before_headers)
        candidate_rows = rows_by_id(candidate_sheet, candidate_headers)
        before_row_count = before_sheet.max_row - 1
        candidate_row_count = candidate_sheet.max_row - 1
        if len(before_rows) != before_row_count or len(candidate_rows) != candidate_row_count:
            return None
        row_delta = 1 if stage.operation == "insert" else -1 if stage.operation == "delete" else 0
        expected_count = before_row_count + row_delta
        if candidate_row_count != expected_count:
            return None
        if stage.operation == "delete":
            if stage.target_id in candidate_rows:
                return None
        elif stage.operation == "insert":
            if candidate_rows.get(stage.target_id) != stage.after:
                return None
        elif stage.target_id not in before_rows or stage.target_id not in candidate_rows:
            return None
        else:
            for column, before_value in before_rows[stage.target_id].items():
                expected = stage.values[column] if column in stage.values else before_value
                if candidate_rows[stage.target_id].get(column) != expected:
                    return None

        unchanged_ids = set(before_rows) & set(candidate_rows) - {stage.target_id}
        if any(before_rows[row_id] != candidate_rows[row_id] for row_id in unchanged_ids):
            return None
        formula_row_ids = unchanged_ids | (
            {stage.target_id} if stage.operation == "update" else set()
        )
        preserved_formula_cells = sum(
            1
            for row_id in formula_row_ids
            for column, before_value in before_rows[row_id].items()
            if isinstance(before_value, str)
            and before_value.startswith("=")
            and candidate_rows[row_id].get(column) == before_value
        )
        return {
            "expected_row_count": expected_count,
            "observed_row_count": candidate_row_count,
            "unchanged_rows_verified": len(unchanged_ids),
            "preserved_formula_cells": preserved_formula_cells,
        }


class WorkbookToolExecutor:
    """Adapter that gives AgentLoop exactly the four permitted tools."""

    def __init__(
        self,
        session: WorkbookSession,
        query_result_callback: Callable[[QueryResult], None] | None = None,
    ) -> None:
        self._session = session
        self._query_result_callback = query_result_callback

    def execute(self, tool_call: ToolCall) -> ToolResult:
        if tool_call.name == "describe_workbook":
            return self._session.describe_workbook()
        if tool_call.name == "query_workbook":
            result = self._session.query_workbook(tool_call.arguments)
            if result.status == "ok" and self._query_result_callback is not None:
                self._query_result_callback(QUERY_RESULT_ADAPTER.validate_python(result.payload))
            return result
        if tool_call.name == "stage_mutation":
            return self._session.stage_mutation(tool_call.arguments)
        if tool_call.name == "commit_mutation":
            return self._session.commit_mutation(tool_call.arguments)
        return ToolResult(status="rejected", payload={"error_code": "unknown_tool"})
