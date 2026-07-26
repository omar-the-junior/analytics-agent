import hashlib
import shutil
from pathlib import Path

from app.agent_loop import AgentLoop, ModelMessage, ToolCall
from app.workbook_session import SOURCES, WorkbookSession, WorkbookToolExecutor
from openpyxl import load_workbook


def test_describe_and_query_preserve_the_supplied_source(tmp_path: Path) -> None:
    source_hash = SOURCES["listings"].read_bytes()
    session = WorkbookSession("listings", tmp_path)

    description = session.describe_workbook()
    result = session.query_workbook({"filters": {"Listing Status": "Active"}, "aggregate": "count"})

    assert description.status == "ok"
    assert description.payload["stable_id"] == "Listing ID"
    assert result.payload["count"] > 0
    assert SOURCES["listings"].read_bytes() == source_hash


def test_session_binds_immutable_source_identity_without_disclosing_paths(
    tmp_path: Path,
) -> None:
    session = WorkbookSession("campaigns", tmp_path)

    source_identity = session.source_hash
    description = session.describe_workbook()

    assert len(source_identity) == 64
    assert description.payload["source_identity"] == source_identity
    assert "path" not in description.payload


def test_session_identity_matches_the_copied_workbook_when_source_changes_during_setup(
    tmp_path: Path, monkeypatch
) -> None:
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"original workbook")
    monkeypatch.setitem(SOURCES, "campaigns", source)
    original_copy = shutil.copy2

    def copy_replaced_source(source_path: Path, destination: Path, *args, **kwargs) -> str:
        source_path.write_bytes(b"replacement workbook")
        return original_copy(source_path, destination, *args, **kwargs)

    monkeypatch.setattr("app.workbook_session.shutil.copy2", copy_replaced_source)

    session = WorkbookSession("campaigns", tmp_path / "session")

    assert session.source_hash == hashlib.sha256(b"replacement workbook").hexdigest()
    assert session.active_path.read_bytes() == b"replacement workbook"


def test_malformed_query_shapes_are_rejected_without_reading_the_workbook(tmp_path: Path) -> None:
    session = WorkbookSession("listings", tmp_path)

    non_object = session.query_workbook([])
    nested_filter = session.query_workbook({"filters": {"City": ["Aurora"]}})
    boolean_limit = session.query_workbook({"limit": True})

    assert non_object.payload == {"error_code": "invalid_query"}
    assert nested_filter.payload == {"error_code": "invalid_query"}
    assert boolean_limit.payload == {"error_code": "invalid_query"}


def test_ambiguous_city_and_unknown_field_fail_safely(tmp_path: Path) -> None:
    session = WorkbookSession("listings", tmp_path)

    ambiguous = session.query_workbook({"filters": {"City": "Aurora"}})
    unknown = session.query_workbook({"filters": {"drop table": "x"}})

    assert ambiguous.status == "needs_clarification"
    assert ambiguous.payload["error_code"] == "ambiguous_city_scope"
    assert unknown.status == "rejected"


def test_stage_requires_exact_commit_and_verifies_artifact(tmp_path: Path) -> None:
    source_bytes = SOURCES["listings"].read_bytes()
    session = WorkbookSession("listings", tmp_path)
    original = session.query_workbook({"filters": {"Listing ID": "LST-5001"}}).payload["rows"][0]
    before_stage = session.active_path.read_bytes()
    version_before_stage = session.describe_workbook().payload["version"]

    staged = session.stage_mutation(
        {"operation": "update", "target_id": "LST-5001", "values": {"List Price": 351001}}
    )
    assert staged.payload["status"] == "confirmation_required"
    assert staged.payload["before"] == {"List Price": 351000}
    assert staged.payload["after"] == {"List Price": 351001}
    assert session.active_path.read_bytes() == before_stage
    assert session.describe_workbook().payload["version"] == version_before_stage

    denied = session.commit_mutation({"stage_id": "not-the-stage"})
    committed = session.commit_mutation({"stage_id": staged.payload["stage_id"]})
    changed = session.query_workbook({"filters": {"Listing ID": "LST-5001"}}).payload["rows"][0]
    assert denied.payload["error_code"] == "confirmation_required"
    assert committed.payload["verified"] is True
    assert original["List Price"] == 351000
    assert changed["List Price"] == 351001
    assert SOURCES["listings"].read_bytes() == source_bytes


def test_commit_reports_verified_artifact_postconditions_and_preserves_formulas(
    tmp_path: Path,
) -> None:
    session = WorkbookSession("listings", tmp_path)
    before_version = session.describe_workbook().payload["version"]
    before_rows = session.query_workbook({"aggregate": "count"}).payload["count"]
    workbook = load_workbook(session.active_path)
    sheet = workbook.active
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(2, headers["Bedrooms"]).value = "=1+2"
    workbook.save(session.active_path)

    staged = session.stage_mutation(
        {"operation": "update", "target_id": "LST-5001", "values": {"List Price": 351001}}
    )
    committed = session.commit_mutation({"stage_id": staged.payload["stage_id"]})

    assert committed.status == "ok"
    assert committed.payload["version"] == before_version + 1
    assert committed.payload["verification"] == {
        "expected_row_count": before_rows,
        "observed_row_count": before_rows,
        "unchanged_rows_verified": before_rows - 1,
        "preserved_formula_cells": 1,
    }
    output = load_workbook(tmp_path / committed.payload["artifact"], data_only=False)
    assert output.active.cell(2, headers["Bedrooms"]).value == "=1+2"


def test_stale_or_missing_authorization_does_not_advance_the_session_version(
    tmp_path: Path,
) -> None:
    session = WorkbookSession("campaigns", tmp_path)
    staged = session.stage_mutation(
        {"operation": "update", "target_id": "CMP-8001", "values": {"Budget Allocated": 1}}
    )
    version_before = session.describe_workbook().payload["version"]
    missing = session.commit_mutation({})
    assert missing.payload == {"error_code": "confirmation_required"}
    assert session.describe_workbook().payload["version"] == version_before

    workbook = load_workbook(session.active_path)
    workbook.active.cell(2, 1).value = "unrelated concurrent edit"
    workbook.save(session.active_path)
    stale = session.commit_mutation({"stage_id": staged.payload["stage_id"]})

    assert stale.payload == {"error_code": "stale_stage"}
    assert session.describe_workbook().payload["version"] == version_before


def test_stage_rejects_malformed_or_unsupported_mutations(tmp_path: Path) -> None:
    session = WorkbookSession("campaigns", tmp_path)

    malformed = session.stage_mutation([])
    unsupported = session.stage_mutation(
        {"operation": "rename", "target_id": "CMP-1001", "values": {}}
    )
    empty_target = session.stage_mutation(
        {"operation": "update", "target_id": "", "values": {"Budget Allocated": 1}}
    )
    whitespace_target = session.stage_mutation(
        {"operation": "update", "target_id": " ", "values": {"Budget Allocated": 1}}
    )
    changed_stable_id = session.stage_mutation(
        {"operation": "update", "target_id": "CMP-8001", "values": {"Campaign ID": "CMP-9000"}}
    )
    listings = WorkbookSession("listings", tmp_path / "listings")
    active_listing = listings.query_workbook({"filters": {"Listing Status": "Active"}}).payload[
        "rows"
    ][0]
    sold_without_sale_price = listings.stage_mutation(
        {
            "operation": "update",
            "target_id": active_listing["Listing ID"],
            "values": {"Listing Status": "Sold"},
        }
    )

    assert malformed.payload == {"error_code": "invalid_mutation"}
    assert unsupported.payload == {"error_code": "invalid_mutation"}
    assert empty_target.payload == {"error_code": "invalid_mutation"}
    assert whitespace_target.payload == {"error_code": "invalid_mutation"}
    assert changed_stable_id.payload == {"error_code": "stable_id_immutable"}
    assert sold_without_sale_price.payload == {"error_code": "missing_sale_price"}


def test_insert_and_delete_are_reviewed_before_they_change_a_session_workbook(
    tmp_path: Path,
) -> None:
    source_bytes = SOURCES["campaigns"].read_bytes()
    insert_session = WorkbookSession("campaigns", tmp_path / "insert")
    template = insert_session.query_workbook(
        {"filters": {"Campaign ID": "CMP-8001"}}
    ).payload["rows"][0]
    inserted_row = {**template, "Campaign ID": "CMP-9999", "Campaign Name": "New campaign"}
    before_insert = insert_session.active_path.read_bytes()
    insert_row_count = insert_session.query_workbook({"aggregate": "count"}).payload["count"]

    inserted = insert_session.stage_mutation(
        {"operation": "insert", "target_id": "CMP-9999", "values": inserted_row}
    )

    assert inserted.payload["status"] == "confirmation_required"
    assert inserted.payload["before"] is None
    assert inserted.payload["after"]["Campaign ID"] == "CMP-9999"
    assert insert_session.active_path.read_bytes() == before_insert
    assert SOURCES["campaigns"].read_bytes() == source_bytes

    committed_insert = insert_session.commit_mutation({"stage_id": inserted.payload["stage_id"]})
    assert committed_insert.payload["verified"] is True
    assert committed_insert.payload["verification"] == {
        "expected_row_count": insert_row_count + 1,
        "observed_row_count": insert_row_count + 1,
        "unchanged_rows_verified": insert_row_count,
        "preserved_formula_cells": 0,
    }
    insert_artifact = load_workbook(tmp_path / "insert" / committed_insert.payload["artifact"])
    assert any(cell.value == "CMP-9999" for cell in insert_artifact.active["A"])
    assert insert_session.query_workbook({"filters": {"Campaign ID": "CMP-9999"}}).payload[
        "count"
    ] == 1
    assert SOURCES["campaigns"].read_bytes() == source_bytes

    delete_session = WorkbookSession("campaigns", tmp_path / "delete")
    delete_row_count = delete_session.query_workbook({"aggregate": "count"}).payload["count"]
    deleted = delete_session.stage_mutation(
        {"operation": "delete", "target_id": "CMP-8001", "values": {}}
    )

    assert deleted.payload["before"]["Campaign ID"] == "CMP-8001"
    assert deleted.payload["after"] is None
    assert delete_session.query_workbook({"filters": {"Campaign ID": "CMP-8001"}}).payload[
        "count"
    ] == 1

    committed_delete = delete_session.commit_mutation({"stage_id": deleted.payload["stage_id"]})
    assert committed_delete.payload["verified"] is True
    assert committed_delete.payload["verification"] == {
        "expected_row_count": delete_row_count - 1,
        "observed_row_count": delete_row_count - 1,
        "unchanged_rows_verified": delete_row_count - 1,
        "preserved_formula_cells": 0,
    }
    delete_artifact = load_workbook(tmp_path / "delete" / committed_delete.payload["artifact"])
    assert all(cell.value != "CMP-8001" for cell in delete_artifact.active["A"])
    assert delete_session.query_workbook({"filters": {"Campaign ID": "CMP-8001"}}).payload[
        "count"
    ] == 0
    assert SOURCES["campaigns"].read_bytes() == source_bytes


def test_tool_executor_only_exposes_the_session_contract(tmp_path: Path) -> None:
    executor = WorkbookToolExecutor(WorkbookSession("campaigns", tmp_path))

    result = executor.execute(ToolCall(name="query_workbook", arguments={"aggregate": "count"}))
    rejected = executor.execute(ToolCall(name="shell", arguments={}))

    assert result.status == "ok"
    assert rejected.status == "rejected"


def test_agent_loop_uses_real_read_executor_without_expanding_its_authority(
    tmp_path: Path,
) -> None:
    class ScriptedModel:
        def __init__(self) -> None:
            self.requests: list[list[ModelMessage]] = []
            self.responses = [
                '{"kind":"tool_batch","tool_calls":[{"name":"query_workbook",'
                '"arguments":{"aggregate":"count"}}]}',
                '{"kind":"final","answer":"Count returned."}',
            ]

        def complete(self, messages: list[ModelMessage]) -> str:
            self.requests.append(messages)
            return self.responses.pop(0)

    model = ScriptedModel()
    run = AgentLoop(model, WorkbookToolExecutor(WorkbookSession("campaigns", tmp_path))).run(
        "Count campaigns.", "Use workbook tools."
    )

    assert run.status == "completed"
    assert run.answer == "Count returned."
    assert '"tool":"query_workbook"' in model.requests[1][-1].content
